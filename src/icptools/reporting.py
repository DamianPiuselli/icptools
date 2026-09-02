from pathlib import Path
from typing import Dict, Any, List, Optional, Union
import pandas as pd
from .models import Batch, SampleMatrix, SampleType
from .pipeline import Processor

def generate_reports(batch: Batch, include_raw: bool = False) -> Dict[str, pd.DataFrame]:
    """
    Generates multiple DataFrames, grouping samples by their SampleMatrix and Prep Group.
    Returns a dictionary mapping a report name (e.g., 'Liquid_None', 'Solid_Digestion_A') to its DataFrame.
    """
    reports = {}
    
    # Group samples by (Matrix, Prep Group)
    groups: Dict[tuple, list] = {}
    for sample in batch.samples:
        key = (sample.matrix, sample.prep_group)
        if key not in groups:
            groups[key] = []
        groups[key].append(sample)
        
    for (matrix, prep_group), samples in groups.items():
        data = []
        # Determine the appropriate unit string based on matrix
        final_unit = "ug/L" if matrix == SampleMatrix.LIQUID else "mg/kg"
        
        for sample in samples:
            row: Dict[str, Any] = {
                "Sample ID": sample.sample_id,
                "Type": sample.sample_type.value,
                "Prep Group": sample.prep_group if sample.prep_group else "N/A"
            }
            
            if matrix == SampleMatrix.SOLID:
                row["Mass (g)"] = sample.mass
                row["% Humidity"] = sample.percent_humidity
                row["Dry Mass (g)"] = sample.dry_mass
                row["Final Volume (mL)"] = sample.final_volume
            
            row["Dilution Factor"] = sample.dilution_factor
            
            # Use requested_analytes if available, otherwise all non-ISTD analytes
            if sample.requested_analytes:
                analytes_to_report = sample.requested_analytes
            else:
                analytes_to_report = [
                    name for name, a in batch.analytes.items() if not a.is_internal_standard
                ]
            
            for analyte_name in analytes_to_report:
                if include_raw:
                    row[f"{analyte_name} [Raw CPS]"] = sample.raw_intensities.get(analyte_name, None)
                    row[f"{analyte_name} [IS Ratio]"] = sample.is_corrected_intensities.get(analyte_name, None)
                    
                row[f"{analyte_name} [Instr Conc (ug/L)]"] = sample.instrument_concentrations.get(analyte_name, None)
                
                # Check if we successfully calculated a final concentration for this analyte
                if analyte_name in sample.final_concentrations:
                    row[f"{analyte_name} [Final Conc ({final_unit})]"] = sample.final_concentrations[analyte_name]
                else:
                    row[f"{analyte_name} [Final Conc ({final_unit})]"] = None
                
            data.append(row)
            
        group_name = f"{matrix.value}_{prep_group if prep_group else 'Direct'}"
        reports[group_name] = pd.DataFrame(data)
        
    return reports

def get_results_dataframe(batch: Batch, include_raw: bool = False) -> pd.DataFrame:
    """
    Convenience method to return a single combined DataFrame for the entire batch.
    Will lack specific unit annotations in the headers because matrix can vary.
    """
    data = []
    
    for sample in batch.samples:
        row: Dict[str, Any] = {
            "Sample ID": sample.sample_id,
            "Type": sample.sample_type.value,
            "Matrix": sample.matrix.value,
            "Prep Group": sample.prep_group,
            "Mass (g)": sample.mass,
            "% Humidity": sample.percent_humidity,
            "Dry Mass (g)": sample.dry_mass,
            "Final Volume (mL)": sample.final_volume,
            "Dilution Factor": sample.dilution_factor
        }
        
        for analyte_name, analyte in batch.analytes.items():
            if analyte.is_internal_standard and not include_raw:
                continue
            if include_raw:
                row[f"{analyte_name} [Raw CPS]"] = sample.raw_intensities.get(analyte_name, None)
                row[f"{analyte_name} [IS Ratio]"] = sample.is_corrected_intensities.get(analyte_name, None)
                
            row[f"{analyte_name} [Instr Conc]"] = sample.instrument_concentrations.get(analyte_name, None)
            row[f"{analyte_name} [Final Conc]"] = sample.final_concentrations.get(analyte_name, None)
            
        data.append(row)
        
    return pd.DataFrame(data)

def get_calibration_summary(processor: Processor) -> pd.DataFrame:
    """
    Returns a DataFrame summarizing the calibration curves for the batch.
    """
    data = []
    for analyte_name, model in processor.calibration_models.items():
        data.append({
            "Analyte": analyte_name,
            "Weighting": model.weighting,
            "Slope": model.slope,
            "Intercept": model.intercept,
            "R-squared": model.r_squared,
            "Fit Successful": model.fit_successful
        })
        
    return pd.DataFrame(data)

def export_excel_formula_report(
    batch: Batch,
    output_path: Union[str, Path],
    instrument_loqs: Optional[Dict[str, float]] = None,
    analytes: Optional[List[str]] = None,
    samples: Optional[List[str]] = None,
    sheet_title: str = "Analytical_Report",
    report_title: Optional[str] = None,
) -> Path:
    """
    Exports a single-sheet, multi-element Excel workbook populated with live
    Excel formulas (SLOPE, INTERCEPT, RSQ, inverse prediction, dilution,
    blank subtraction, prep math, and dynamic Sample LoQ evaluation).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    out_file = Path(output_path)
    out_file.parent.mkdir(parents=True, exist_ok=True)

    loq_dict = instrument_loqs or {}

    # 1. Determine target analytes
    if analytes:
        target_analytes = [a for a in analytes if a in batch.analytes]
    else:
        # Check if samples specify requested_analytes
        req_set = []
        for s in batch.samples:
            if s.requested_analytes:
                for a in s.requested_analytes:
                    if a in batch.analytes and a not in req_set:
                        req_set.append(a)
        if req_set:
            target_analytes = req_set
        else:
            target_analytes = [
                name for name, a in batch.analytes.items() if not a.is_internal_standard
            ]

    # 2. Determine calibration standards
    # Collect all standards relevant to target analytes
    cal_samples = []
    cal_sample_ids = set()
    for s in batch.samples:
        has_known = any(a in s.known_concentrations for a in target_analytes)
        is_std_type = s.sample_type in (SampleType.STANDARD, SampleType.BLANK)
        if has_known and is_std_type:
            cal_samples.append(s)
            cal_sample_ids.add(s.sample_id)

    # 3. Determine analytical unknown / client samples
    if samples:
        target_samples = [batch.get_sample(sid) for sid in samples if batch.get_sample(sid) is not None]
    else:
        target_samples = [s for s in batch.samples if s.sample_id not in cal_sample_ids]

    # Check if solid matrix is present among target samples
    has_solid = any(s.matrix == SampleMatrix.SOLID for s in target_samples)

    # Initialize Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel tab title max 31 chars
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    font_title = Font(name="Calibri", size=13, bold=True, color="1F497D")
    font_sub = Font(name="Calibri", size=10, italic=True, color="595959")
    font_sec = Font(name="Calibri", size=11, bold=True, color="1F497D")
    font_head = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_reg = Font(name="Calibri", size=10)
    font_rep = Font(name="Calibri", size=10, bold=True, color="1F497D")

    fill_cal_head = PatternFill("solid", fgColor="4F81BD")
    fill_param = PatternFill("solid", fgColor="DCE6F1")
    fill_sample_head = PatternFill("solid", fgColor="243F60")
    fill_rep_head = PatternFill("solid", fgColor="366092")
    fill_blank_row = PatternFill("solid", fgColor="F2F2F2")

    thin_border = Side(border_style="thin", color="D9D9D9")
    med_border = Side(border_style="medium", color="4F81BD")
    med_dark = Side(border_style="medium", color="243F60")
    cell_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    cal_head_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=med_border)
    sample_head_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=med_dark)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # -------------------------------------------------------------------------
    # SECTION 1: HEADER
    # -------------------------------------------------------------------------
    title_text = report_title or f"ICP-MS Analytical Report - {batch.name}"
    ws.cell(row=1, column=1, value=title_text).font = font_title
    ws.cell(row=2, column=1, value=f"Batch: {batch.name} | Processed with icptools").font = font_sub

    # -------------------------------------------------------------------------
    # SECTION 2: CALIBRATION STANDARDS & REGRESSION TABLE
    # -------------------------------------------------------------------------
    ws.cell(row=4, column=1, value="1. CALIBRATION STANDARDS & REGRESSION PARAMETERS").font = font_sec

    cal_headers = ["Standard ID", "Level"]
    for a in target_analytes:
        is_name = batch.analyte_to_is.get(a, "None")
        cal_headers.append(f"{a} Nominal (ug/L)")
        cal_headers.append(f"{a} Ratio ({a}/{is_name})")

    # Write calibration headers
    cal_head_row = 5
    for col_idx, h in enumerate(cal_headers, 1):
        c = ws.cell(row=cal_head_row, column=col_idx, value=h)
        c.font = font_head
        c.fill = fill_cal_head
        c.border = cal_head_border
        c.alignment = align_center

    # Write calibration rows
    cal_start_row = 6
    analyte_std_rows: Dict[str, List[int]] = {a: [] for a in target_analytes}

    for idx, s in enumerate(cal_samples):
        r = cal_start_row + idx
        ws.cell(row=r, column=1, value=s.sample_id).alignment = align_left
        ws.cell(row=r, column=2, value=s.level if s.level is not None else "").alignment = align_center

        for a_idx, a in enumerate(target_analytes):
            nom_col = 3 + a_idx * 2
            ratio_col = 4 + a_idx * 2

            if a in s.known_concentrations:
                analyte_std_rows[a].append(r)
                nom_conc = s.known_concentrations[a]
                ratio_val = s.is_corrected_intensities.get(a, 0.0)

                c_nom = ws.cell(row=r, column=nom_col, value=nom_conc)
                c_nom.number_format = "#,##0.00"
                c_nom.alignment = align_right

                c_rat = ws.cell(row=r, column=ratio_col, value=ratio_val)
                c_rat.number_format = "0.000000"
                c_rat.alignment = align_right
            else:
                # Blank if this standard does not calibrate analyte 'a'
                ws.cell(row=r, column=nom_col, value=None)
                ws.cell(row=r, column=ratio_col, value=None)

        for col_i in range(1, len(cal_headers) + 1):
            ws.cell(row=r, column=col_i).font = font_reg
            ws.cell(row=r, column=col_i).border = cell_border

    cal_end_row = cal_start_row + len(cal_samples) - 1

    # Calibration Parameter Rows (Slope, Intercept, R2, LoQ)
    param_labels = [
        ("Slope (m)", "0.000000"),
        ("Intercept (b)", "0.000000"),
        ("R-squared (R²)", "0.00000"),
        ("Instrument LoQ (ug/L)", "#,##0.00"),
    ]

    slope_cells: Dict[str, str] = {}
    intercept_cells: Dict[str, str] = {}
    rsq_cells: Dict[str, str] = {}
    loq_cells: Dict[str, str] = {}

    for p_idx, (lbl, num_fmt) in enumerate(param_labels):
        r = cal_end_row + 1 + p_idx
        c_lbl = ws.cell(row=r, column=1, value=lbl)
        c_lbl.font = font_bold
        c_lbl.fill = fill_param
        c_lbl.border = cell_border
        c_lbl.alignment = align_left

        ws.cell(row=r, column=2, value="").fill = fill_param
        ws.cell(row=r, column=2).border = cell_border

        for a_idx, a in enumerate(target_analytes):
            nom_col_letter = get_column_letter(3 + a_idx * 2)
            rat_col_letter = get_column_letter(4 + a_idx * 2)

            # Determine row range for analyte 'a'
            rows_for_a = analyte_std_rows[a]
            if rows_for_a:
                a_start_r = min(rows_for_a)
                a_end_r = max(rows_for_a)
            else:
                a_start_r = cal_start_row
                a_end_r = cal_end_row

            c_val = ws.cell(row=r, column=4 + a_idx * 2)
            c_blank = ws.cell(row=r, column=3 + a_idx * 2)
            c_blank.fill = fill_param
            c_blank.border = cell_border

            if p_idx == 0:  # Slope
                c_val.value = f"=SLOPE({rat_col_letter}{a_start_r}:{rat_col_letter}{a_end_r}, {nom_col_letter}{a_start_r}:{nom_col_letter}{a_end_r})"
                slope_cells[a] = f"${rat_col_letter}${r}"
            elif p_idx == 1:  # Intercept
                c_val.value = f"=INTERCEPT({rat_col_letter}{a_start_r}:{rat_col_letter}{a_end_r}, {nom_col_letter}{a_start_r}:{nom_col_letter}{a_end_r})"
                intercept_cells[a] = f"${rat_col_letter}${r}"
            elif p_idx == 2:  # R2
                c_val.value = f"=RSQ({rat_col_letter}{a_start_r}:{rat_col_letter}{a_end_r}, {nom_col_letter}{a_start_r}:{nom_col_letter}{a_end_r})"
                rsq_cells[a] = f"${rat_col_letter}${r}"
            elif p_idx == 3:  # LoQ
                c_val.value = float(loq_dict.get(a, 0.1))
                loq_cells[a] = f"${rat_col_letter}${r}"

            c_val.font = font_bold
            c_val.fill = fill_param
            c_val.border = cell_border
            c_val.number_format = num_fmt
            c_val.alignment = align_right

    # -------------------------------------------------------------------------
    # SECTION 3: SAMPLES PROCESSING TABLE
    # -------------------------------------------------------------------------
    sample_start_row = cal_end_row + 7
    ws.cell(row=sample_start_row - 2, column=1, value="2. SAMPLE MEASUREMENTS & DATA PROCESSING").font = font_sec

    # Build Sample Headers (Metadata)
    s_headers = ["Sample ID", "Type", "Matrix", "Dilution"]
    if has_solid:
        s_headers.extend(["Mass (g)", "% Humidity", "Vol (mL)"])

    meta_len = len(s_headers)
    dil_col_letter = get_column_letter(4)
    mass_col_letter = get_column_letter(5) if has_solid else None
    hum_col_letter = get_column_letter(6) if has_solid else None
    vol_col_letter = get_column_letter(7) if has_solid else None

    # Determine which analytes have method blanks in this sample set
    mblk_rows: Dict[tuple, int] = {}  # prep_group -> row_number

    # Pre-calculate column offsets for each analyte
    # Columns per analyte: [Ratio, Instr Conc, (optional Blank-Corr), Sample LoQ, Final Conc, Reported]
    analyte_col_info: Dict[str, Dict[str, int]] = {}
    curr_col = meta_len + 1

    unit_str = "mg/kg" if has_solid else "ug/L"
    has_method_blank = any(s.sample_type == SampleType.METHOD_BLANK for s in target_samples)

    for a in target_analytes:
        info = {}
        is_name = batch.analyte_to_is.get(a, "IS")

        s_headers.append(f"{a} Ratio ({a}/{is_name})")
        info["ratio"] = curr_col
        curr_col += 1

        s_headers.append(f"{a} Instr Conc (ug/L)")
        info["instr_conc"] = curr_col
        curr_col += 1

        if has_method_blank:
            s_headers.append(f"{a} Blank-Corr (ug/L)")
            info["blank_corr"] = curr_col
            curr_col += 1

        s_headers.append(f"{a} Sample LoQ ({unit_str})")
        info["sample_loq"] = curr_col
        curr_col += 1

        s_headers.append(f"{a} Final Conc ({unit_str})")
        info["final_conc"] = curr_col
        curr_col += 1

        analyte_col_info[a] = info

    # Write Sample Headers
    head_row_num = sample_start_row - 1
    for col_idx, h in enumerate(s_headers, 1):
        c = ws.cell(row=head_row_num, column=col_idx, value=h)
        c.font = font_head
        c.fill = fill_sample_head
        c.border = sample_head_border
        c.alignment = align_center

    # First pass: map method blank rows
    for s_idx, s in enumerate(target_samples):
        r = sample_start_row + s_idx
        if s.sample_type == SampleType.METHOD_BLANK:
            mblk_rows[s.prep_group] = r

    # Second pass: write sample rows and formulas
    for s_idx, s in enumerate(target_samples):
        r = sample_start_row + s_idx
        is_blank_sample = (s.sample_type == SampleType.METHOD_BLANK)

        # Metadata
        ws.cell(row=r, column=1, value=s.sample_id).alignment = align_left
        ws.cell(row=r, column=2, value=s.sample_type.value).alignment = align_center
        ws.cell(row=r, column=3, value=s.matrix.value).alignment = align_center
        c_dil = ws.cell(row=r, column=4, value=s.dilution_factor)
        c_dil.number_format = "#,##0.0"
        c_dil.alignment = align_right

        if has_solid:
            c_mass = ws.cell(row=r, column=5, value=s.mass)
            c_mass.number_format = "#,##0.0000"
            c_mass.alignment = align_right

            c_hum = ws.cell(row=r, column=6, value=s.percent_humidity)
            c_hum.number_format = "0.0"
            c_hum.alignment = align_right

            # Display volume in mL: if stored in L (<= 1.0), convert to mL for readability
            vol_val = s.final_volume * 1000.0 if s.final_volume <= 1.0 else s.final_volume
            c_vol = ws.cell(row=r, column=7, value=vol_val)
            c_vol.number_format = "#,##0.0"
            c_vol.alignment = align_right

        for a in target_analytes:
            cols = analyte_col_info[a]

            ratio_col_let = get_column_letter(cols["ratio"])
            instr_col_let = get_column_letter(cols["instr_conc"])
            final_col_let = get_column_letter(cols["final_conc"])
            loq_col_let = get_column_letter(cols["sample_loq"])

            # 1. Ratio (directly written value as requested)
            ratio_val = s.is_corrected_intensities.get(a, 0.0)
            c_rat = ws.cell(row=r, column=cols["ratio"], value=ratio_val)
            c_rat.number_format = "0.000000"
            c_rat.alignment = align_right

            # 2. Instr Conc formula: =(Ratio - Intercept) / Slope
            int_ref = intercept_cells[a]
            slp_ref = slope_cells[a]
            c_instr = ws.cell(row=r, column=cols["instr_conc"], value=f"=({ratio_col_let}{r}-{int_ref})/{slp_ref}")
            c_instr.number_format = "#,##0.00"
            c_instr.alignment = align_right

            # 3. Blank-Correction (if applicable)
            active_conc_let = instr_col_let
            if "blank_corr" in cols:
                blank_r = mblk_rows.get(s.prep_group)
                bcorr_col_let = get_column_letter(cols["blank_corr"])
                if blank_r and not is_blank_sample:
                    ws.cell(row=r, column=cols["blank_corr"], value=f"={instr_col_let}{r}-${instr_col_let}${blank_r}")
                else:
                    ws.cell(row=r, column=cols["blank_corr"], value=f"={instr_col_let}{r}")
                c_bcorr = ws.cell(row=r, column=cols["blank_corr"])
                c_bcorr.number_format = "#,##0.00"
                c_bcorr.alignment = align_right
                active_conc_let = bcorr_col_let

            # 4. Sample LoQ formula (dry weight basis)
            inst_loq_ref = loq_cells[a]
            if s.matrix == SampleMatrix.SOLID and has_solid:
                # Solid Method LoQ: =(Instr_LoQ * (Vol_mL / 1000) * Dilution) / (Mass_g * (1 - Hum/100))
                ws.cell(
                    row=r,
                    column=cols["sample_loq"],
                    value=f"=({inst_loq_ref}*({vol_col_letter}{r}/1000)*{dil_col_letter}{r})/({mass_col_letter}{r}*(1-{hum_col_letter}{r}/100))"
                )
            else:
                # Liquid Method LoQ: =Instr_LoQ * Dilution
                ws.cell(row=r, column=cols["sample_loq"], value=f"={inst_loq_ref}*{dil_col_letter}{r}")
            c_sloq = ws.cell(row=r, column=cols["sample_loq"])
            c_sloq.number_format = "#,##0.00"
            c_sloq.alignment = align_right

            # 5. Final Conc formula (dry weight basis)
            if s.matrix == SampleMatrix.SOLID and has_solid:
                # Solid Final Conc: =((Conc) * (Vol_mL / 1000) * Dilution) / (Mass_g * (1 - Hum/100))
                ws.cell(
                    row=r,
                    column=cols["final_conc"],
                    value=f"=(({active_conc_let}{r})*({vol_col_letter}{r}/1000)*{dil_col_letter}{r})/({mass_col_letter}{r}*(1-{hum_col_letter}{r}/100))"
                )
            else:
                # Liquid Final Conc: =Conc * Dilution
                ws.cell(row=r, column=cols["final_conc"], value=f"={active_conc_let}{r}*{dil_col_letter}{r}")
            c_fconc = ws.cell(row=r, column=cols["final_conc"])
            c_fconc.number_format = "#,##0.00"
            c_fconc.alignment = align_right

        # Apply row borders and styling
        for col_i in range(1, len(s_headers) + 1):
            cell_i = ws.cell(row=r, column=col_i)
            cell_i.border = cell_border
            cell_i.font = font_reg
            if is_blank_sample:
                cell_i.fill = fill_blank_row

    sample_end_row = sample_start_row + len(target_samples) - 1

    # -------------------------------------------------------------------------
    # SECTION 4: CLIENT REPORT SUMMARY TABLE (CLEAN, COMPACT, COPY-PASTE READY)
    # -------------------------------------------------------------------------
    client_start_row = sample_end_row + 4
    ws.cell(row=client_start_row - 2, column=1, value="3. FINAL CLIENT REPORT (REPORTED RESULTS)").font = font_sec

    # Client report headers: Sample ID followed by element names without mass number
    client_headers = ["Sample ID"]
    for a in target_analytes:
        elem_name = batch.analytes[a].element if (a in batch.analytes and batch.analytes[a].element) else a
        client_headers.append(f"{elem_name} ({unit_str})")

    # Write Client Report Headers
    cl_head_row_num = client_start_row - 1
    for col_idx, h in enumerate(client_headers, 1):
        c = ws.cell(row=cl_head_row_num, column=col_idx, value=h)
        c.font = font_head
        c.fill = fill_rep_head
        c.border = sample_head_border
        c.alignment = align_center

    # Filter for client samples only (exclude laboratory method blanks and QCs)
    client_samples = [
        (s_idx, s) for s_idx, s in enumerate(target_samples)
        if s.sample_type not in (SampleType.METHOD_BLANK, SampleType.QC)
    ]

    # Write Client Report Rows (referencing Section 2 calculation cells)
    for cl_idx, (s_idx, s) in enumerate(client_samples):
        r_client = client_start_row + cl_idx
        r_proc = sample_start_row + s_idx

        c_id = ws.cell(row=r_client, column=1, value=s.sample_id)
        c_id.alignment = align_left
        c_id.font = font_bold
        c_id.border = cell_border

        for a_idx, a in enumerate(target_analytes):
            cols = analyte_col_info[a]
            final_col_let = get_column_letter(cols["final_conc"])
            loq_col_let = get_column_letter(cols["sample_loq"])

            # Formula referencing Section 2's Final Conc and Sample LoQ
            c_val = ws.cell(
                row=r_client,
                column=2 + a_idx,
                value=f'=IF({final_col_let}{r_proc}<{loq_col_let}{r_proc}, "< " & TEXT({loq_col_let}{r_proc}, "0.00"), ROUND({final_col_let}{r_proc}, 2))'
            )
            c_val.font = font_rep
            c_val.alignment = align_right
            c_val.border = cell_border

    # -------------------------------------------------------------------------
    # Auto-adjust column widths
    # -------------------------------------------------------------------------
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or "")
            if val_str.startswith("="):
                val_str = "1234567.89"  # approximate formula output length
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 13)

    wb.save(out_file)
    return out_file
